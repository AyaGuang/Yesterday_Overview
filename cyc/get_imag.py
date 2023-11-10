import requests
from bs4 import BeautifulSoup
import re
from openpyxl import Workbook

# 发送GET请求，获取搜索结果页面的HTML内容
def get_search_results_html():
    #设计html标头
    url = "https://you.ctrip.com/sight/longnan2424/4383341.html"
    header = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36 Edg/119.0.0.0",
    }
    #遍历所有网页，生成总的html文件
    response = requests.get(url,headers=header)
    html=response.text
    return html

# 解析HTML，提取图片链接
def get_info(html):
    info=[]
    soup = BeautifulSoup(html, "html.parser")
    info_elements = soup.find_all("div", class_="commentItem")
    for info_element in info_elements:
        comment_element=info_element.find("div",class_="commentDetail")
        images_element=info_element.find("div",class_="commentImgList")
        image_links=[]
        if images_element:
            img_elements=images_element.find_all("img")
            for img_element in img_elements:
                if img_element:
                    img_src = img_element.get("src")
                    image_links.append(img_src)
        info.append([comment_element.get_text(strip=True),image_links])
    return info

# 将结果保存到excel文件中
def save_excel(info):
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = '评论'
    ws.cell(row=1, column=2).value = '图片链接'
    offset=0
    for i in range(len(info)):
        ws.cell(row=i + 2 + offset, column=1).value = info[i][0]
        for j in range(len(info[i][1])):
            ws.cell(row=i + 2 + j + offset, column=2).value = info[i][1][j]
        offset+=len(info[i][1])
    wb.save('out.xlsx')

# 主函数
def main():
    print("------开始爬取，等待一段时间------")
    search_results_html = get_search_results_html()
    print(search_results_html)
    image_links = get_info(search_results_html)
    save_excel(image_links)
    print("------爬取结束，结果保存至excel------")

main()